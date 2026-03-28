import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Employee Import"

# Define headers
headers = ['employeeNumber', 'firstName', 'lastName', 'email', 'phone', 'password', 
           'designation', 'department', 'planId', 'coverageStartDate', 'coverageEndDate', 'cnic', 'dob']

# Add headers with formatting
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Sample data
sample_data = [
    ['EMP001', 'Ahmed', 'Khan', 'ahmed.khan@company.com', '+92-300-1234567', 'SecurePass123!', 'Senior Developer', 'Engineering', 'plan-001', '2026-01-01', '2027-01-01', '12345-1234567-1', '1990-05-15'],
    ['EMP002', 'Fatima', 'Ali', 'fatima.ali@company.com', '03211234567', 'TechPass456!', 'Project Manager', 'Engineering', 'plan-001', '2026-01-01', '2027-01-01', '12345-1234567-2', '1988-08-22'],
    ['EMP003', 'Hassan', 'Ahmed', 'hassan.ahmed@company.com', '+92-300-2345678', 'DevPass789!', 'Full Stack Developer', 'Engineering', 'plan-002', '2026-01-15', '2027-01-15', '12345-1234567-3', '1992-03-10'],
    ['EMP004', 'Ayesha', 'Hassan', 'ayesha.hassan@company.com', '03121234568', 'FinPass123!', 'Accountant', 'Finance', 'plan-001', '2026-02-01', '2027-02-01', '12345-1234567-4', '1994-11-30'],
    ['EMP005', 'Ali', 'Raza', 'ali.raza@company.com', '+92-300-3456789', 'HRPass456!', 'HR Manager', 'People', 'plan-003', '2026-02-15', '2027-02-15', '12345-1234567-5', '1991-07-18'],
    ['EMP006', 'Zainab', 'Mohammad', 'zainab.mohammad@company.com', '03001234569', 'SalesPass789!', 'Sales Executive', 'Sales', 'plan-002', '2026-03-01', '2027-03-01', '12345-1234567-6', '1995-09-25'],
    ['EMP007', 'Muhammad', 'Hasan', 'muhammad.hasan@company.com', '+92-300-4567890', 'ITPass123!', 'IT Support', 'IT', 'plan-001', '2026-03-15', '2027-03-15', '12345-1234567-7', '1993-12-05'],
    ['EMP008', 'Nida', 'Khan', 'nida.khan@company.com', '03211234570', 'DesignPass456!', 'UI/UX Designer', 'Design', 'plan-002', '2026-04-01', '2027-04-01', '12345-1234567-8', '1996-02-14'],
    ['EMP009', 'Omar', 'Ali', 'omar.ali@company.com', '+92-300-5678901', 'LogisPass789!', 'Logistics Coordinator', 'Logistics', 'plan-001', '2026-04-15', '2027-04-15', '12345-1234567-9', '1989-06-20'],
    ['EMP010', 'Samina', 'Hassan', 'samina.hassan@company.com', '03121234571', 'ProdPass123!', 'Production Manager', 'Production', 'plan-003', '2026-05-01', '2027-05-01', '12345-1234567-10', '1987-10-12'],
]

# Add data rows
for row_idx, row_data in enumerate(sample_data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.alignment = Alignment(horizontal="left", vertical="center")

# Adjust column widths
column_widths = [15, 15, 15, 25, 18, 18, 20, 15, 12, 16, 16, 18, 12]
for idx, width in enumerate(column_widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

# Save the workbook
wb.save('employee-import-sample.xlsx')
print("Excel file created successfully: employee-import-sample.xlsx")
